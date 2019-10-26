# -*- coding: UTF-8 -*-

#telegram_bot_learn_english_phrases

import openpyxl
import requests

def send_telegram(text: str, channel_id):
    #'https': 'socks5://150.242.254.52:1080'
    proxies = {
        
        'https': 'socks5://190.26.19.250:9999',
        'https': 'socks5://50.62.57.97:1736',
        'https': 'socks5://51.68.142.150:9300',
        'https': 'socks5://166.62.91.254:14718',
        'https': 'socks5://132.148.128.250:35172',
        'https': 'socks5://104.238.74.55:4273',
        'https': 'socks5://128.14.23.137:1080',
        'https': 'socks5://195.9.134.86:8888'
        
    }    
    token = "token"
    url = "https://api.telegram.org/bot"
    #channel_id = "@sergo_english_bot"
    #channel_id=845468268 # Sergo telega
    #channel_id=947637323 # Nastya telega
    #channel_id=386235031 # artem telega
    #channel_id=928686647 # nik telega
    #channel_id=215151271 # english bot telega
    url += token
    method = url + "/sendMessage"

    r = requests.post(method, data={
         "chat_id": channel_id,
         "text": text
          }, proxies=proxies)

    if r.status_code != 200:
        raise Exception("Something wrong, sent text error...")


def xlref(row, column, zero_indexed=True):
    '''
    xlref - Simple conversion of row, column to an excel string format
    >>> xlref(0,0)
    'A1'
    >>> xlref(0,26)
    'AA1'    
    '''
    if zero_indexed:
        row += 1
        column += 1
    return openpyxl.utils.get_column_letter(column) + str(row)

def selection_phrase(filepath):
    wb = openpyxl.load_workbook(filepath)
    number_of_phrases = 5
    fill_color = openpyxl.styles.PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00') #YELLOW
    ws = wb.active    
    index_to_last_words = int(ws.cell(row=1, column=1).value)
    for i in range(index_to_last_words, index_to_last_words + number_of_phrases):
        ws[(xlref(i,1))].fill = fill_color #B1B2
        eng_phrase = (ws.cell(row=i+1, column=2).value)
        ws[(xlref(i,2))].fill = fill_color #C1C2
        rus_phrase = (ws.cell(row=i+1, column=3).value)
        column_letter= (openpyxl.utils.get_column_letter(2) + str(i+1)) # index to Last words
        ws.cell(row=1, column=1,value= column_letter.replace("B", "")) # insert A1:cell index to Last words
        txt = (eng_phrase + ' => ' + rus_phrase)
        send_telegram(txt, channel_id=947637323)# Nastya telega
        send_telegram(txt, channel_id=845468268)# Sergo telega
        
    wb.save(filepath)    
    

if __name__ == '__main__':
    filepath =(r'D:\Win_XP\Рабочий стол\Admin txt\Script\Python\telegram_bot_english_learn\learn_engl_phrases.xlsx')
    #filepath =(r'I:\testdata\engl_words.xlsx')
    selection_phrase(filepath)
    












'''

# печатаем список листов
sheets = wb.sheetnames
for sheet in sheets:
    print(sheet)

# получаем активный лист
sheet = wb.active

# печатаем значение ячейки A1
print(sheet['A1'].value)
# печатаем значение ячейки B1
print(sheet['B1'].value)
cell = sheet['B2']
cell_Eng = sheet.cell(row = 2, column = 2)
cell_Rus = sheet.cell(row = 2, column = 3)
print(cell_Eng.value,"-",cell_Rus.value)


for row in sheet['B1':'C1']:
    string = ''
    for cell in row:
        string = string + str(cell.value) + ' '
    print(string)
'''

'''
for row in sheet.iter_rows(values_only=True):
...         print(row)

'''